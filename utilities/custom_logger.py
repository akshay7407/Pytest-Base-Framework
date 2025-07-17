import inspect
import logging
from datetime import datetime
import allure
import os

time = datetime.now()


def custom_logger():
    log_name = inspect.stack()[1][3]
    logger = logging.getLogger(log_name)
    logger.setLevel(logging.DEBUG)
    file_handler = logging.FileHandler(
        os.path.normpath(os.getcwd() + os.sep ).rstrip("/utilities/") + "/reports/test_reports.log", mode='a')
    print(file_handler)
    file_handler.setLevel(logging.DEBUG)
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s : %(message)s',
                                  datefmt='%d/%m/%y %I:%M:%S %p %A')
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)
    return logger


from openpyxl import load_workbook
from datetime import datetime
 
def add_data_to_excel():
    """
    Adds values to a new or existing sheet in an Excel file under specified headers,
    including a 'Generated Date' column with today's date.
    
    Args:
        file_path (str): Path to the Excel file.
        sheet_name (str): Name of the sheet to add or update.
        headers (list of str): List of column headers.
        values (list of any): List of values corresponding to the headers.
    """
       # Headers and values to write
    headers = ["Name", "Age", "Grade","Gernrated Date"]
    name = "John Doe"
    age = 21
    grade = "A"
    values = [name, age, grade]

    sheet_name ="MetaData"

    file_path ="D:\\Pytest-Base-Framework\\Sample Data.xlsx"
    
    try:
        workbook = load_workbook(file_path)
    except FileNotFoundError:
        raise FileNotFoundError(f"Excel file not found at {file_path}")
 
    # Add or access the sheet
    if sheet_name not in workbook.sheetnames:
        worksheet = workbook.create_sheet(sheet_name)
    else:
        worksheet = workbook[sheet_name]
 
    # Validate headers
    # existing_headers = [cell.value for cell in worksheet[1]]  # Read the first row
    # if existing_headers != headers:
    #     raise ValueError(f"Headers in the sheet do not match the provided headers: {existing_headers} != {headers}")
 
    # Append today's date to values
    today_date = datetime.now().strftime("%Y-%m-%d")
    values.append(today_date)
 
    max_row = worksheet.max_row
    for i , (headers,values) in enumerate(zip(headers,values),start=1):
        worksheet.cell(row=max_row + i, column=1).value = headers  # Write headers in column A
        worksheet.cell(row=max_row + i, column=2).value = values  # Write values in column B

 
    # Save changes
    workbook.save(file_path)

result = add_data_to_excel() 
print(result)